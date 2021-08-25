from openpyxl import load_workbook
from icecream import ic

file_name = 'transfer_fee.xlsx'

# Open Excel File
wb = load_workbook(file_name,
                   read_only=True,
                   data_only=True) # True: 수식이 아닌 수식의 결과값, False: 수식

# Get Sheet List
# ws_list = wb.get_sheet_names()
# for ws in ws_list:
#     ic(ws)

# Select Worksheet
ws = wb['항공요금']

# Get Cell Value
# cell_A1 = ws['A1']
# ic(cell_A1)
# ic(cell_A1.value)

# Get All 'A' Column
# index = 1
# while True:
#     cell = 'A%d' % index
#     cell_value = ws[cell].value
#
#     if cell_value is None:
#         break
#
#     ic('%s: %s' % (cell, cell_value))
#     index += 1

# Get All 1st Row # A1, B1, C1, D1, ....
# index = 0
# while True:
#     cell = '%s1' % (chr(ord('A') + index)) # ord('A') -> ASCII(42) 42+1=43 -> 'B'
#     cell_value = ws[cell].value
#
#     if cell_value is None:
#         break
#     ic('%s: %s' % (cell, cell_value))
#     index += 1

# Get row data
# index = 1
# # for row in ws.rows[1:]: # TypeError: 'generator' object is not subscriptable
# for row in ws.rows:
#     # if index == 1:
#     #     index += 1
#     #     continue
#     print('%s line: ' % (index), end='')
#     for i in range(len(row)):
#         print('%s ' % (row[i].value), end='')
#     index += 1
#     print()

index = 1
for row in ws.iter_rows(min_row=2):
    print('%s line: ' % (index), end='')
    for i in range(len(row)):
        print('%s ' % (row[i].value), end='')
    index += 1
    print()