from openpyxl import Workbook
from datetime import datetime
from openpyxl.drawing.image import Image

# Create Excel Object
wb = Workbook()

file_name = 'new_excel.xlsx'

# Get Active Worksheet
ws = wb.active

# Set Sheet Names
ws.title = 'my_data'

# Set Cell Value
ws['A1'] = 'HelloWorld'

# Insert Row
row = [1,2,3,4,5]
ws.append(row)

# Use Function
# ws['A3'] = '=SUM(1+3)' # =SUM(A1:B1)
ws['A3'] = '=SUM(A2:B2)' # =SUM(A1:B1) Cell 값이 숫자로 입력되야만 가능

# Insert DateTime
ws['A4'] = datetime.now()

# Merge Cell
ws['F1'] = 'Hello'
ws['G1'] = 'World'
ws.merge_cells('F1:H1') # 결과는 F1값만 남음
# ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=8)
ws.unmerge_cells('F1:H1')
# ws.unmerge_cells(start_row=1, start_column=6, end_row=1, end_column=8)

# Insert Image
img = Image('car.jpg')
ws.add_image(img, 'A10')

wb.save(filename=file_name)
